import pandas as pd
import os
from openpyxl import Workbook

# Input paths
op_1_path = "op_1.csv"  # Seating arrangement output
ip_4_path = "ip_4.csv"  # Roll and name mapping

# Output folder
output_folder = "attendance_sheets"

# Ensure output folder exists
os.makedirs(output_folder, exist_ok=True)

# Load seating arrangement data and roll-name mapping
op_1 = pd.read_csv(op_1_path, sep=";")
ip_4 = pd.read_csv(ip_4_path, sep=";")

# Create a dictionary for roll-to-name mapping
roll_name_mapping = dict(zip(ip_4["Roll"], ip_4["Name"]))

# Generate attendance sheets for each date
for date in op_1["Date"].unique():
    date_data = op_1[op_1["Date"] == date]
    date_formatted = pd.to_datetime(date).strftime("%d_%m_%Y")
    date_output_path = os.path.join(output_folder, f"Attendance_{date_formatted}.xlsx")

    # Track whether any sheet is created
    sheets_created = False

    # Create a new workbook
    workbook = Workbook()
    default_sheet = workbook.active
    default_sheet.title = "Placeholder"  # Temporary placeholder sheet

    # Process Morning slot
    for course_code in date_data["Morning"].unique():
        if course_code == "NO EXAM":
            continue
        course_data = date_data[date_data["Morning"] == course_code]
        for room in course_data["Room"].unique():
            room_data = course_data[course_data["Room"] == room]

            # Prepare the attendance sheet data
            attendance_data = []
            for roll_list in room_data["Roll_list"]:
                for roll in roll_list.split(";"):
                    name = roll_name_mapping.get(roll, "Unknown")
                    attendance_data.append([roll, name, ""])  # Roll, Name, Empty column for signature

            # Convert to DataFrame
            attendance_df = pd.DataFrame(attendance_data, columns=["Roll No.", "Name", "Signature"])

            # Add placeholder rows for invigilator and TA signatures
            placeholders = pd.DataFrame({"Roll No.": [""] * 5, "Name": [""] * 5, "Signature": [""] * 5})
            attendance_df = pd.concat([attendance_df, placeholders], ignore_index=True)

            # Generate sheet name
            sheet_name = f"{date_formatted}_{course_code}_Room{room}_Morning"
            sheet_name = sheet_name[:31]  # Limit sheet name to 31 characters
            workbook.create_sheet(title=sheet_name)
            ws = workbook[sheet_name]

            # Write column headers
            ws.append(["Roll No.", "Name", "Signature"])

            # Write attendance data
            for row in attendance_df.itertuples(index=False):
                ws.append(row)
            sheets_created = True

    # Process Evening slot
    for course_code in date_data["Evening"].unique():
        if course_code == "NO EXAM":
            continue
        course_data = date_data[date_data["Evening"] == course_code]
        for room in course_data["Room"].unique():
            room_data = course_data[course_data["Room"] == room]

            # Prepare the attendance sheet data
            attendance_data = []
            for roll_list in room_data["Roll_list"]:
                for roll in roll_list.split(";"):
                    name = roll_name_mapping.get(roll, "Unknown")
                    attendance_data.append([roll, name, ""])  # Roll, Name, Empty column for signature

            # Convert to DataFrame
            attendance_df = pd.DataFrame(attendance_data, columns=["Roll No.", "Name", "Signature"])

            # Add placeholder rows for invigilator and TA signatures
            placeholders = pd.DataFrame({"Roll No.": [""] * 5, "Name": [""] * 5, "Signature": [""] * 5})
            attendance_df = pd.concat([attendance_df, placeholders], ignore_index=True)

            # Generate sheet name
            sheet_name = f"{date_formatted}_{course_code}_Room{room}_Evening"
            sheet_name = sheet_name[:31]  # Limit sheet name to 31 characters
            workbook.create_sheet(title=sheet_name)
            ws = workbook[sheet_name]

            # Write column headers
            ws.append(["Roll No.", "Name", "Signature"])

            # Write attendance data
            for row in attendance_df.itertuples(index=False):
                ws.append(row)
            sheets_created = True

    # If no sheets were created, skip the file
    if not sheets_created:
        print(f"No valid data for date {date}. Skipping file creation.")
        continue

    # Remove the placeholder sheet if other sheets exist
    if "Placeholder" in workbook.sheetnames:
        del workbook["Placeholder"]

    # Save the workbook
    workbook.save(date_output_path)

print("Attendance sheets generated successfully!")
